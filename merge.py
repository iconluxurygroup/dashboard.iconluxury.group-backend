from pydantic import BaseModel
from fastapi.middleware.cors import CORSMiddleware
from typing import List
from fastapi import FastAPI, UploadFile, Form, HTTPException
import pyodbc
import uuid
import os,re
import pandas as pd
import shutil
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
import boto3
from typing import Optional
import logging


# Lightweight job model for initial list
class JobSummary(BaseModel):
    id: int
    inputFile: str
    fileEnd: str | None
    user: str
    rec: int
    img: int

# Full job details model
class ResultItem(BaseModel):
    resultId: int
    entryId: int
    imageUrl: str
    imageDesc: str | None
    imageSource: str | None
    createTime: str | None
    imageUrlThumbnail: str | None
    sortOrder: int
    imageIsFashion: int
    aiCaption: str | None
    aiJson: str | None
    aiLabel: str | None

class RecordItem(BaseModel):
    entryId: int
    fileId: int
    excelRowId: int
    productModel: str | None
    productBrand: str | None
    createTime: str | None
    step1: str | None
    step2: str | None
    step3: str | None
    step4: str | None
    completeTime: str | None
    productColor: str | None
    productCategory: str | None
    excelRowImageRef: str | None; 

class JobDetails(BaseModel):
    id: int
    inputFile: str
    imageStart: str | None
    fileStart: str | None
    fileEnd: str | None
    resultFile: str | None
    fileLocationUrl: str | None
    logFileUrl: str | None
    user: str
    rec: int
    img: int
    apiUsed: str
    imageEnd: str | None
    results: list[ResultItem]
    records: list[RecordItem]

# Model for domain aggregation (matches React's DomainAggregation)
class DomainAggregation(BaseModel):
    domain: str
    totalResults: int
    positiveSortOrderCount: int

# CORS middleware to allow requests from your frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Adjust to specific origins (e.g., "http://localhost:3000") in production
    allow_credentials=True,
    allow_methods=["GET", "OPTIONS", "POST"],
    allow_headers=["*"],
)

# MSSQL connection settings
DB_CONFIG = {
    "server": "35.172.243.170",
    "database": "luxurymarket_p4",
    "username": "luxurysitescraper",
    "password": "Ftu5675FDG54hjhiuu$",
    "driver": "{ODBC Driver 17 for SQL Server}",
}

# AWS S3 configuration (optional, adjust as needed)
S3_CONFIG = {
    "endpoint": "https://s3.us-east-2.amazonaws.com",  # e.g., "https://nyc3.digitaloceanspaces.com"
    "region": "us-east-2",  # e.g., "us-east-1"
    "access_key": "AKIAZQ3DSIQ5BGLY355N",
    "secret_key": "uB1D2M4/dXz4Z6as1Bpan941b3azRM9N770n1L6Q",
    "bucket_name": "iconluxurygroup-s3",
}

# Set up logging
logging.basicConfig(level=logging.INFO)
default_logger = logging.getLogger(__name__)

# Database connection function
def get_db_connection():
    conn_str = (
        f"DRIVER={DB_CONFIG['driver']};"
        f"SERVER={DB_CONFIG['server']};"
        f"DATABASE={DB_CONFIG['database']};"
        f"UID={DB_CONFIG['username']};"
        f"PWD={DB_CONFIG['password']}"
    )
    return pyodbc.connect(conn_str)

# S3 client setup (optional)
s3_client = boto3.client(
    "s3",
    region_name=S3_CONFIG["region"],
    endpoint_url=S3_CONFIG["endpoint"],
    aws_access_key_id=S3_CONFIG["access_key"],
    aws_secret_access_key=S3_CONFIG["secret_key"],
)
import mimetypes
def upload_to_s3(local_file_path, bucket_name, s3_key):
    try:
        content_type, _ = mimetypes.guess_type(local_file_path)
        if not content_type:
            content_type = 'application/octet-stream'
            default_logger.warning(f"Could not determine Content-Type for {local_file_path}")
        s3_client.upload_file(
            local_file_path,
            bucket_name,
            s3_key,
            ExtraArgs={
                'ACL': 'public-read',
                'ContentType': content_type
            }
        )
        s3_url = f"{S3_CONFIG['endpoint']}/{bucket_name}/{s3_key}"
        default_logger.info(f"Uploaded {local_file_path} to S3: {s3_url} with Content-Type: {content_type}")
        return s3_url
    except Exception as e:
        default_logger.error(f"Failed to upload_column_mapload_column_map {local_file_path} to S3: {e}")
        raise
# Function to extract data and images from Excel
import os
import uuid
import shutil
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader

def extract_data_and_images(file_path, file_id, column_map):
    # Load Excel workbook and sheet
    wb = load_workbook(file_path)
    sheet = wb.active
    image_loader = SheetImageLoader(sheet)

    # Directory for temporary image storage
    extracted_images_dir = os.path.join("temp_files", "extracted_images", file_id)
    os.makedirs(extracted_images_dir, exist_ok=True)

    extracted_data = []
    for row_idx in range(2, sheet.max_row + 1):
        data = {
            'search': sheet[f'{column_map["search"]}{row_idx}'].value,
            'brand': sheet[f'{column_map["brand"]}{row_idx}'].value,
            'ExcelRowImageRef': sheet[f'{column_map["image"]}{row_idx}'].value, 
            'color': (sheet[f'{column_map["color"]}{row_idx}'].value if column_map["color"] else None),
            'category': (sheet[f'{column_map["category"]}{row_idx}'].value if column_map["category"] else None),
        }
        default_logger.info(f"Extracted data for row {row_idx}: {data}")
        # Handle embedded images if no image reference exists
        image_cell = f'{column_map["image"]}{row_idx}'
        if not data['ExcelRowImageRef'] and image_loader.image_in(image_cell):
            img_path = os.path.join(extracted_images_dir, f"image_{file_id}_{image_cell}.png")
            image = image_loader.get(image_cell)
            if image:
                image.save(img_path)
                # Upload to S3
                s3_key = f"images/{file_id}/{os.path.basename(img_path)}"
                s3_url = upload_to_s3(img_path, S3_CONFIG['bucket_name'], s3_key)
                data['ExcelRowImageRef'] = s3_url
            else:
                default_logger.warning(f"No image retrieved from cell {image_cell}")
        else:
            default_logger.info(f"No image found in cell {image_cell}")

        extracted_data.append(data)

    return extracted_data, extracted_images_dir

def validate_column(col: str) -> str:
    if not col or not re.match(r"^[A-Z]+$", col):
        raise HTTPException(
            status_code=400,
            detail=f"Invalid column name: {col}. Must be uppercase letters only (e.g., 'A', 'B', 'AA')."
        )
    return col
def load_payload_db(rows, file_id, column_map, logger=None):
    logger = logger or default_logger
    try:
        file_id = int(file_id)
        conn_str = (
            f"DRIVER={{ODBC Driver 17 for SQL Server}};"
            f"SERVER=35.172.243.170;"
            f"DATABASE=luxurymarket_p4;"
            f"UID=luxurysitescraper;"
            f"PWD={DB_CONFIG['password']}"
        )
        with pyodbc.connect(conn_str) as connection:
            cursor = connection.cursor()
            cursor.execute("SELECT COUNT(*) FROM utb_ImageScraperFiles WHERE ID = ?", (file_id,))
            if cursor.fetchone()[0] == 0:
                raise ValueError(f"FileID {file_id} does not exist in utb_ImageScraperFiles")

            # Create DataFrame from rows
            df = pd.DataFrame(rows).rename(columns={
                column_map['search']: 'ProductModel',
                column_map['brand']: 'ProductBrand',
                column_map.get('color', ''): 'ProductColor',
                column_map.get('category', ''): 'ProductCategory'
            })

            if 'ProductBrand' not in df.columns or df['ProductBrand'].isnull().all():
                logger.warning(f"No valid ProductBrand data found in payload for FileID {file_id}")
                df['ProductBrand'] = df.get('ProductBrand', '')

            # Add FileID and ExcelRowID
            df.insert(0, 'FileID', file_id)
            df.insert(1, 'ExcelRowID', range(1, len(df) + 1))

            # Handle image URL using ExcelRowImageRef (keep as-is)
            if column_map.get('ExcelRowImageRef') in df.columns:
                logger.debug(f"Image URLs loaded in ExcelRowImageRef for FileID {file_id}: {df[column_map['ExcelRowImageRef']].head().tolist()}")
            else:
                logger.warning(f"No ExcelRowImageRef column found in data for FileID {file_id}")
                df['ExcelRowImageRef'] = None  # Add ExcelRowImageRef column with NULL if not present

            logger.debug(f"Inserting data for FileID {file_id}: {df[['ProductBrand', 'ProductModel']].head().to_dict()}")

            # Define expected columns, including ExcelRowImageRef
            expected_cols = ['FileID', 'ExcelRowID', 'ProductModel', 'ProductBrand', 'ProductColor', 'ProductCategory', 'ExcelRowImageRef']

            # Normalize columns
            for col in expected_cols:
                if col not in df.columns:
                    df[col] = None
                elif col in ['ProductColor', 'ProductCategory']:
                    df[col] = pd.to_numeric(df[col], errors='coerce')
                    df[col] = df[col].where(df[col].notna(), None)

            # Insert into database
            for _, row in df.iterrows():
                row_values = [None if pd.isna(val) else val for val in row[expected_cols]]
                cursor.execute(
                    f"INSERT INTO utb_ImageScraperRecords ({', '.join(expected_cols)}) VALUES ({', '.join(['?'] * len(expected_cols))})",
                    tuple(row_values)
                )
            connection.commit()
        logger.info(f"Loaded {len(df)} rows into utb_ImageScraperRecords for FileID: {file_id}")
        return df
    except Exception as e:
        logger.error(f"Error loading payload data: {e}")
        raise
# Insert file metadata into database
def insert_file_db(filename: str, file_url: str, email: Optional[str]) -> int:
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        query = """
            INSERT INTO utb_ImageScraperFiles (FileName, FileLocationUrl, UserEmail, CreateFileStartTime)
            OUTPUT INSERTED.ID
            VALUES (?, ?, ?, GETDATE())
        """
        cursor.execute(query, (filename, file_url, email or 'default@example.com'))
        row = cursor.fetchone()
        if row is None or row[0] is None:
            raise Exception("Insert failed or no identity value returned.")
        file_id = int(row[0])
        conn.commit()
        conn.close()
        return file_id
    except pyodbc.Error as e:
        default_logger.error(f"Database error: {e}")
        raise
    except Exception as e:
        default_logger.error(f"Error in insert_file_db: {e}")
        raise

@app.post("/submitImage")
async def submit_image(
    fileUploadImage: UploadFile,
    imageColumnImage: Optional[str] = Form(None),  # Make it optional
    searchColImage: str = Form(...),
    brandColImage: str = Form(...),
    ColorColImage: Optional[str] = Form(None),
    CategoryColImage: Optional[str] = Form(None),
    sendToEmail: Optional[str] = Form(None),
):
    temp_dir = None
    extracted_images_dir = None
    try:
        # Generate unique file ID
        file_id = str(uuid.uuid4())
        default_logger.info(f"Processing file for FileID: {file_id}")

        # Save Excel file locally
        temp_dir = os.path.join("temp_files", "images", file_id)
        os.makedirs(temp_dir, exist_ok=True)
        uploaded_file_path = os.path.join(temp_dir, fileUploadImage.filename)
        with open(uploaded_file_path, "wb") as buffer:
            shutil.copyfileobj(fileUploadImage.file, buffer)

        # Upload Excel file to S3
        s3_key_excel = f"uploads/{file_id}/{fileUploadImage.filename}"
        file_url = upload_to_s3(uploaded_file_path, S3_CONFIG['bucket_name'], s3_key_excel)

        # Validate and map columns
        extract_column_map = {
            'search': validate_column(searchColImage),
            'brand': validate_column(brandColImage),
            'image': validate_column(imageColumnImage) if imageColumnImage else None,  # Only validate if provided
            'color': validate_column(ColorColImage) if ColorColImage else None,
            'category': validate_column(CategoryColImage) if CategoryColImage else None,
        }

        # Extract data and images
        extracted_data, extracted_images_dir = extract_data_and_images(uploaded_file_path, file_id, extract_column_map)

        # Insert file metadata into database
        file_id_db = insert_file_db(fileUploadImage.filename, file_url, sendToEmail)

        # Insert extracted data into database
        load_column_map = {
            'search': 'search',
            'brand': 'brand',
            'color': 'color' if extract_column_map['color'] else '',
            'category': 'category' if extract_column_map['category'] else '',
            'ExcelRowImageRef': 'ExcelRowImageRef'
        }
        load_payload_db(extracted_data, file_id_db, load_column_map, default_logger)

        # Return success response
        return {
            "success": True,
            "fileUrl": file_url,
            "message": "File uploaded and processed successfully",
            "file_id": file_id_db,
        }

    except Exception as e:
        default_logger.error(f"Error processing file: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Server error: {str(e)}")
    finally:
        # Clean up temporary files
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)
        if extracted_images_dir and os.path.exists(extracted_images_dir):
            shutil.rmtree(extracted_images_dir, ignore_errors=True)

def extract_data_and_images(file_path, file_id, column_map):
    # Load Excel workbook and sheet
    wb = load_workbook(file_path)
    sheet = wb.active
    image_loader = SheetImageLoader(sheet) if column_map['image'] else None  # Only load if image column is provided

    # Directory for temporary image storage (only create if we might need it)
    extracted_images_dir = None
    if column_map['image']:
        extracted_images_dir = os.path.join("temp_files", "extracted_images", file_id)
        os.makedirs(extracted_images_dir, exist_ok=True)

    extracted_data = []
    for row_idx in range(2, sheet.max_row + 1):  # Start from row 2 assuming row 1 is headers
        image_ref = None
        if column_map['image']:
            image_cell = f'{column_map["image"]}{row_idx}'
            image_ref = sheet[image_cell].value  # Get the value in the image column if provided

        data = {
            'search': sheet[f'{column_map["search"]}{row_idx}'].value,
            'brand': sheet[f'{column_map["brand"]}{row_idx}'].value,
            'ExcelRowImageRef': image_ref if image_ref else None,
            'color': (sheet[f'{column_map["color"]}{row_idx}'].value if column_map["color"] else None),
            'category': (sheet[f'{column_map["category"]}{row_idx}'].value if column_map["category"] else None),
        }
        default_logger.info(f"Extracted data for row {row_idx}: {data}")

        # Only scan for embedded images if image column is provided and no reference exists
        if column_map['image'] and not data['ExcelRowImageRef'] and image_loader and image_loader.image_in(image_cell):
            img_path = os.path.join(extracted_images_dir, f"image_{file_id}_{image_cell}.png")
            image = image_loader.get(image_cell)
            if image:
                image.save(img_path)
                # Upload to S3
                s3_key = f"images/{file_id}/{os.path.basename(img_path)}"
                s3_url = upload_to_s3(img_path, S3_CONFIG['bucket_name'], s3_key)
                data['ExcelRowImageRef'] = s3_url
                default_logger.info(f"Extracted and uploaded image from {image_cell} to {s3_url}")
            else:
                default_logger.warning(f"No image retrieved from cell {image_cell}")
        elif column_map['image'] and not data['ExcelRowImageRef']:
            default_logger.info(f"No image reference or embedded image found in cell {image_cell}")
        elif column_map['image'] and data['ExcelRowImageRef']:
            default_logger.info(f"Using provided image reference from {image_cell}: {data['ExcelRowImageRef']}")
        else:
            default_logger.info(f"No image column provided, skipping image extraction for row {row_idx}")

        extracted_data.append(data)

    return extracted_data, extracted_images_dir
@app.get("/api/scraping-jobs", response_model=list[JobSummary])
async def get_all_jobs(page: int = 1, page_size: int = 10):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        offset = (page - 1) * page_size

        query = """
            SELECT 
                ID, 
                FileName, 
                CreateFileCompleteTime, 
                UserEmail,
                (SELECT COUNT(*) FROM utb_ImageScraperRecords WHERE FileID = utb_ImageScraperFiles.ID) as rec_count,
                (SELECT COUNT(*) FROM utb_ImageScraperResult 
                 WHERE EntryID IN (SELECT EntryID FROM utb_ImageScraperRecords WHERE FileID = utb_ImageScraperFiles.ID)) as img_count
            FROM utb_ImageScraperFiles
            ORDER BY ID DESC
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """
        cursor.execute(query, (offset, page_size))
        rows = cursor.fetchall()

        jobs_data = [
            {
                "id": row.ID,
                "inputFile": row.FileName,
                "fileEnd": row.CreateFileCompleteTime.isoformat() if row.CreateFileCompleteTime else None,
                "user": row.UserEmail or "Unknown",
                "rec": row.rec_count,
                "img": row.img_count,
            }
            for row in rows
        ]

        conn.close()
        return jobs_data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Detailed job endpoint (unchanged)
@app.get("/api/scraping-jobs/{job_id}", response_model=JobDetails)
async def get_job(job_id: int):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        query_files = """
            SELECT ID, FileName, ImageStartTime, CreateFileStartTime, CreateFileCompleteTime,
                   FileLocationURLComplete AS ResultFile, FileLocationUrl, LogFileURL, UserEmail,
                   ImageCompleteTime
            FROM utb_ImageScraperFiles
            WHERE ID = ?
        """
        cursor.execute(query_files, (job_id,))
        row = cursor.fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Job not found")

        query_results = """
            SELECT ResultID, EntryID, ImageUrl, ImageDesc, ImageSource, CreateTime, ImageUrlThumbnail,
                   SortOrder, ImageIsFashion, AiCaption, AiJson, AiLabel
            FROM utb_ImageScraperResult
            WHERE EntryID IN (SELECT EntryID FROM utb_ImageScraperRecords WHERE FileID = ?)
        """
        cursor.execute(query_results, (job_id,))
        results = cursor.fetchall()

        query_records = """
            SELECT EntryID, FileID, ExcelRowID, ProductModel, ProductBrand, CreateTime, Step1, Step2, 
                   Step3, Step4, CompleteTime, ProductColor, ProductCategory,excelRowImageRef
            FROM utb_ImageScraperRecords
            WHERE FileID = ?
        """
        cursor.execute(query_records, (job_id,))
        records = cursor.fetchall()

        job_data = {
            "id": row.ID,
            "inputFile": row.FileName,
            "imageStart": row.ImageStartTime.isoformat() if row.ImageStartTime else None,
            "fileStart": row.CreateFileStartTime.isoformat() if row.CreateFileStartTime else None,
            "fileEnd": row.CreateFileCompleteTime.isoformat() if row.CreateFileCompleteTime else None,
            "resultFile": row.ResultFile,
            "fileLocationUrl": row.FileLocationUrl,
            "logFileUrl": row.LogFileURL,
            "user": row.UserEmail or "Unknown",
            "rec": len(records),
            "img": len(results),
            "apiUsed": "google-serp",
            "imageEnd": row.ImageCompleteTime.isoformat() if row.ImageCompleteTime else None,
            "results": [
                {
                    "resultId": r.ResultID,
                    "entryId": r.EntryID,
                    "imageUrl": r.ImageUrl or "None",
                    "imageDesc": r.ImageDesc,
                    "imageSource": r.ImageSource,
                    "createTime": r.CreateTime.isoformat() if r.CreateTime else None,
                    "imageUrlThumbnail": r.ImageUrlThumbnail or "None",
                    "sortOrder": r.SortOrder or -1,
                    "imageIsFashion": r.ImageIsFashion,
                    "aiCaption": r.AiCaption,
                    "aiJson": r.AiJson,
                    "aiLabel": r.AiLabel,
                } for r in results
            ],
            "records": [
                {
                    "entryId": r.EntryID,
                    "fileId": r.FileID,
                    "excelRowId": r.ExcelRowID,
                    "productModel": r.ProductModel,
                    "productBrand": r.ProductBrand,
                    "createTime": r.CreateTime.isoformat() if r.CreateTime else None,
                    "step1": r.Step1.isoformat() if r.Step1 else None,
                    "step2": r.Step2.isoformat() if r.Step1 else None,
                    "step3": r.Step3.isoformat() if r.Step1 else None,
                    "step4": r.Step4.isoformat() if r.Step1 else None,
                    "completeTime": r.CompleteTime.isoformat() if r.CompleteTime else None,
                    "productColor": r.ProductColor if r.ProductColor else None,
                    "productCategory": r.ProductCategory if r.ProductCategory else None,
                    "excelRowImageRef":r.excelRowImageRef if r.excelRowImageRef else None,
                } for r in records
            ],
        }

        conn.close()
        return job_data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# New endpoint to aggregate all results by domain
@app.get("/api/whitelist-domains", response_model=List[DomainAggregation])
async def get_whitelist_domains():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Fetch all results from utb_ImageScraperResult
        query = """
            SELECT ImageSource, SortOrder
            FROM utb_ImageScraperResult
        """
        cursor.execute(query)
        rows = cursor.fetchall()

        # Aggregate by domain
        domain_data = {}
        for row in rows:
            image_source = row.ImageSource if row.ImageSource else "unknown"
            sort_order = row.SortOrder if row.SortOrder is not None else -1

            # Extract domain from ImageSource URL
            try:
                from urllib.parse import urlparse
                domain = urlparse(image_source).hostname or "unknown"
                domain = domain.replace("www.", "")  # Remove "www." prefix
            except:
                domain = "unknown"

            if domain not in domain_data:
                domain_data[domain] = {"totalResults": 0, "positiveSortOrderCount": 0}
            
            domain_data[domain]["totalResults"] += 1
            if sort_order > 0:
                domain_data[domain]["positiveSortOrderCount"] += 1

        # Convert to list for response
        aggregated_domains = [
            {
                "domain": domain,
                "totalResults": data["totalResults"],
                "positiveSortOrderCount": data["positiveSortOrderCount"],
            }
            for domain, data in domain_data.items()
        ]

        conn.close()
        return aggregated_domains
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)