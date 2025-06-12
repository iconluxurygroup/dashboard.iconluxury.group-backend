from pydantic import BaseModel
from fastapi.middleware.cors import CORSMiddleware
from typing import List, Dict, Optional, Tuple
from fastapi import FastAPI, UploadFile, Form, HTTPException
import pyodbc
import uuid
import os
import re
import pandas as pd
import shutil
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
import boto3
import requests
import json
import logging
import urllib.parse
import mimetypes
from datetime import datetime
from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
from typing import List, Optional
import pyodbc
from datetime import datetime
from config import VERSION
from email_utils import send_message_email
# Assuming this code is part of the existing FastAPI app
# If it's a separate module, ensure the app is imported or instantiated
# Initialize FastAPI app
app = FastAPI(title="iconluxury.group", version=VERSION)

# Lightweight job model for initial list
class JobSummary(BaseModel):
    id: int
    inputFile: str
    fileEnd: str | None
    user: str
    rec: int
    img: int

# Full job details model
class ResultItem(BaseModel):
    resultId: int
    entryId: int
    imageUrl: str
    imageDesc: str | None
    imageSource: str | None
    createTime: str | None
    imageUrlThumbnail: str | None
    sortOrder: int
    imageIsFashion: int
    aiCaption: str | None
    aiJson: str | None
    aiLabel: str | None

class RecordItem(BaseModel):
    entryId: int
    fileId: int
    excelRowId: int
    productModel: str | None
    productBrand: str | None
    createTime: str | None
    step1: str | None
    step2: str | None
    step3: str | None
    step4: str | None
    completeTime: str | None
    productColor: str | None
    productCategory: str | None
    excelRowImageRef: str | None

class JobDetails(BaseModel):
    id: int
    inputFile: str
    imageStart: str | None
    fileStart: str | None
    fileEnd: str | None
    resultFile: str | None
    fileLocationUrl: str | None
    logFileUrl: str | None
    user: str
    rec: int
    img: int
    apiUsed: str
    imageEnd: str | None
    results: list[ResultItem]
    records: list[RecordItem]

# Model for domain aggregation
class DomainAggregation(BaseModel):
    domain: str
    totalResults: int
    positiveSortOrderCount: int
# Model for job progress
class JobProgress(BaseModel):
    fileId: int
    totalRecords: int
    step1Completed: int
    step1Progress: float
    step2Completed: int
    step2Progress: float
    step3Completed: int
    step3Progress: float
    step4Completed: int
    step4Progress: float


# Model for reference data
class ReferenceData(BaseModel):
    data: Dict[str, str]

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["GET", "OPTIONS", "POST"],
    allow_headers=["*"],
)

# MSSQL connection settings
DB_CONFIG = {
    "server": "35.172.243.170",
    "database": "luxurymarket_p4",
    "username": "luxurysitescraper",
    "password": "Ftu5675FDG54hjhiuu$",
    "driver": "{ODBC Driver 17 for SQL Server}",
}

# AWS S3 and Cloudflare R2 configuration
S3_CONFIG = {
    "endpoint": "https://s3.us-east-2.amazonaws.com",
    "region": "us-east-2",
    "access_key": "AKIA2CUNLEV6V627SWI7",
    "secret_key": "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",
    "bucket_name": "iconluxurygroup",
    "r2_endpoint": "https://aa2f6aae69e7fb4bd8e2cd4311c411cb.r2.cloudflarestorage.com",
    "r2_access_key": "8b5a4a988c474205e0172eab5479d6f2",
    "r2_secret_key": "8ff719bbf2946c1b6a81fcf2121e1a41604a0b6f2890f308871b381e98a8d725",
    "r2_account_id": "aa2f6aae69e7fb4bd8e2cd4311c411cb",
    "r2_bucket_name": "iconluxurygroup",
    "r2_custom_domain": "https://iconluxury.group",
}

# Set up logging
logging.basicConfig(level=logging.INFO)
default_logger = logging.getLogger(__name__)

# Database connection function
def get_db_connection():
    conn_str = (
        f"DRIVER={DB_CONFIG['driver']};"
        f"SERVER={DB_CONFIG['server']};"
        f"DATABASE={DB_CONFIG['database']};"
        f"UID={DB_CONFIG['username']};"
        f"PWD={DB_CONFIG['password']}"
    )
    return pyodbc.connect(conn_str)

# S3 and R2 upload function
def get_s3_client(service='s3', logger=None, file_id=None):
    logger = logger or default_logger
    if logger == default_logger and file_id:
        logger.info(f"Setup logger for get_s3_client, FileID: {file_id}")
    
    try:
        logger.info(f"Creating {service.upper()} client")
        if service == 'r2':
            client = boto3.client(
                "s3",
                region_name='auto',
                endpoint_url=S3_CONFIG['r2_endpoint'],
                aws_access_key_id=S3_CONFIG['r2_access_key'],
                aws_secret_access_key=S3_CONFIG['r2_secret_key']
            )
        else:
            client = boto3.client(
                "s3",
                region_name=S3_CONFIG['region'],
                endpoint_url=S3_CONFIG['endpoint'],
                aws_access_key_id=S3_CONFIG['access_key'],
                aws_secret_access_key=S3_CONFIG['secret_key']
            )
        logger.info(f"{service.upper()} client created successfully")
        return client
    except Exception as e:
        logger.error(f"Error creating {service.upper()} client: {e}", exc_info=True)
        raise

def double_encode_plus(filename, logger=None):
    logger = logger or default_logger
    logger.debug(f"Encoding filename: {filename}")
    first_pass = filename.replace('+', '%2B')
    second_pass = urllib.parse.quote(first_pass)
    logger.debug(f"Double-encoded filename: {second_pass}")
    return second_pass

def upload_to_s3(local_file_path, bucket_name, s3_key, r2_bucket_name=None, logger=None, file_id=None):
    logger = logger or default_logger
    if logger == default_logger and file_id:
        logger.info(f"Setup logger for upload_to_s3, FileID: {file_id}")
    
    result_urls = {}
    
    # Determine Content-Type
    content_type, _ = mimetypes.guess_type(local_file_path)
    if not content_type:
        content_type = 'application/octet-stream'
        logger.warning(f"Could not determine Content-Type for {local_file_path}")
    
    # Upload to AWS S3
    try:
        s3_client = get_s3_client(service='s3', logger=logger, file_id=file_id)
        logger.info(f"Uploading {local_file_path} to S3: {bucket_name}/{s3_key}")
        s3_client.upload_file(
            local_file_path,
            bucket_name,
            s3_key,
            ExtraArgs={
                'ACL': 'public-read',
                'ContentType': content_type
            }
        )
        double_encoded_key = double_encode_plus(s3_key, logger=logger)
        s3_url = f"https://{bucket_name}.s3.{S3_CONFIG['region']}.amazonaws.com/{double_encoded_key}"
        logger.info(f"Uploaded {local_file_path} to S3: {s3_url} with Content-Type: {content_type}")
        result_urls['s3'] = s3_url
    except Exception as e:
        logger.error(f"Failed to upload {local_file_path} to S3: {e}")
        raise
    
    # Upload to Cloudflare R2 (if r2_bucket_name is provided)
    if r2_bucket_name:
        try:
            r2_client = get_s3_client(service='r2', logger=logger, file_id=file_id)
            logger.info(f"Uploading {local_file_path} to R2: {r2_bucket_name}/{s3_key}")
            r2_client.upload_file(
                local_file_path,
                r2_bucket_name,
                s3_key,
                ExtraArgs={
                    'ACL': 'public-read',
                    'ContentType': content_type
                }
            )
            double_encoded_key = double_encode_plus(s3_key, logger=logger)
            r2_url = f"{S3_CONFIG['r2_custom_domain']}/{double_encoded_key}"
            logger.info(f"Uploaded {local_file_path} to R2: {r2_url} with Content-Type: {content_type}")
            result_urls['r2'] = r2_url
        except Exception as e:
            logger.error(f"Failed to upload {local_file_path} to R2: {e}")
            raise
    
    return result_urls


import aiohttp
from fastapi import HTTPException
from datetime import datetime

async def send_file_details_email(
    to_email: str,
    file_id: int,
    filename: str,
    s3_url: str,
    r2_url: str | None,
    record_count: int,
    nikoffer_count: int,
    user_email: str | None
) -> bool:
    try:
        subject = f"File Upload Notification - File ID: {file_id}"
        
        # Construct the restart job URL
        restart_job_url = f"https://icon7-8080.iconluxury.today/api/v4/restart-search-all/{file_id}"
        
        # Construct the email message with file details and restart job link
        message = (
            "File Upload Notification\n"
            f"File ID: {file_id}\n"
            f"File Name: {filename}\n"
            f"S3 URL: {s3_url}\n"
            f"R2 URL: {r2_url or 'Not available'}\n"
            f"Upload Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"Records Processed: {record_count}\n"
            f"Nikoffer Records Processed: {nikoffer_count}\n"
            f"Uploaded By: {user_email or 'Unknown'}\n"
            f"Environment: iconluxury.group\n"
            f"Version: {VERSION}\n"
        )
        
        default_logger.info(f"Preparing to send email to {to_email} with subject: {subject}")
        
        # Send the POST request to restart the search job
        async with aiohttp.ClientSession() as session:
            async with session.post(
                restart_job_url,
                headers={"accept": "application/json"}
            ) as response:
                if response.status == 200:
                    default_logger.info(f"Successfully triggered restart search job for file ID {file_id}")
                else:
                    default_logger.error(f"Failed to trigger restart search job for file ID {file_id}: {response.status}")
                    # Optionally, you could raise an exception or continue without failing the email
                    # raise HTTPException(status_code=500, detail=f"Failed to trigger restart job: {response.status}")

        # Send the email using the provided send_message_email function
        success = await send_message_email(
            to_emails=to_email,
            subject=subject,
            message=message,
            logger=default_logger
        )
        
        if success:
            default_logger.info(f"Email sent successfully to {to_email}")
            return True
        else:
            default_logger.error(f"Failed to send email to {to_email}")
            return False
            
    except Exception as e:
        default_logger.error(f"Error sending file details email to {to_email}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to send email: {str(e)}")


def validate_column(col: str) -> str:
    if not col or not re.match(r"^[A-Z]+$", col):
        raise HTTPException(
            status_code=400,
            detail=f"Invalid column name: {col}. Must be uppercase letters only (e.g., 'A', 'B', 'AA')."
        )
    return col

def load_payload_db(rows, file_id, column_map, logger=None):
    logger = logger or default_logger
    try:
        file_id = int(file_id)
        with get_db_connection() as connection:
            cursor = connection.cursor()
            cursor.execute("SELECT COUNT(*) FROM utb_ImageScraperFiles WHERE ID = ?", (file_id,))
            if cursor.fetchone()[0] == 0:
                raise ValueError(f"FileID {file_id} does not exist in utb_ImageScraperFiles")

            df = pd.DataFrame(rows)
            logger.debug(f"Raw DataFrame (rows={len(df)}): {df.to_dict(orient='records')}")

            # Rename columns to match database fields
            rename_dict = {
                'search': 'ProductModel',
                'brand': 'ProductBrand',
                'color': 'ProductColor',
                'category': 'ProductCategory',
                'ExcelRowImageRef': 'ExcelRowImageRef'
            }
            df = df.rename(columns=rename_dict)
            logger.debug(f"DataFrame after renaming: {df.to_dict(orient='records')}")

            # Apply manual brand if specified
            if column_map['brand'] == 'MANUAL':
                manual_brand_value = column_map.get('manualBrand', '')
                if not manual_brand_value:
                    logger.warning("brandColImage is 'MANUAL' but manualBrand is empty or None")
                df['ProductBrand'] = manual_brand_value
                logger.debug(f"Applied manual brand: '{manual_brand_value}' to all rows")

            # Add required columns
            df['FileID'] = file_id
            df['ExcelRowID'] = range(1, len(df) + 1)

            logger.debug(f"DataFrame after adding FileID and ExcelRowID: {df.to_dict(orient='records')}")

            # Define expected columns
            expected_cols = ['FileID', 'ExcelRowID', 'ProductModel', 'ProductBrand', 'ProductColor', 'ProductCategory', 'ExcelRowImageRef']

            # Ensure all expected columns exist
            for col in expected_cols:
                if col not in df.columns:
                    df[col] = None
                df[col] = df[col].where(df[col].notna(), None)

            logger.debug(f"Final DataFrame before DB insert: {df.head().to_dict(orient='records')}")

            # Insert rows
            rows_inserted = 0
            for idx, row in df.iterrows():
                try:
                    row_values = [row.get(col, None) for col in expected_cols]
                    cursor.execute(
                        f"INSERT INTO utb_ImageScraperRecords ({', '.join(expected_cols)}) VALUES ({', '.join(['?'] * len(expected_cols))})",
                        tuple(row_values)
                    )
                    rows_inserted += 1
                except Exception as e:
                    logger.error(f"Error inserting row {idx + 1}: {e}")

            connection.commit()
            logger.info(f"Committed {rows_inserted} rows into utb_ImageScraperRecords for FileID: {file_id}")

            # Verify insertion
            cursor.execute(
                "SELECT FileID, ExcelRowID, ProductModel, ProductBrand, ProductColor, ProductCategory, ExcelRowImageRef "
                "FROM utb_ImageScraperRecords WHERE FileID = ? ORDER BY ExcelRowID",
                (file_id,)
            )
            inserted_rows = cursor.fetchall()
            logger.debug(f"All data from DB for FileID {file_id}: {inserted_rows}")

        logger.info(f"Loaded {len(df)} rows into utb_ImageScraperRecords for FileID: {file_id}")
        return df
    except Exception as e:
        logger.error(f"Error loading payload data: {e}")
        raise


def insert_file_db(filename: str, file_url: str, email: Optional[str], header_index: int, file_type: int ,logger=None) -> int:
    logger = logger or default_logger
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        query = """
            INSERT INTO utb_ImageScraperFiles (FileName, FileLocationUrl, UserEmail, UserHeaderIndex,FileTypeID, CreateFileStartTime)
            OUTPUT INSERTED.ID
            VALUES (?, ?, ?, ?,?, GETDATE())
        """
        cursor.execute(query, (filename, file_url, email or 'nik@accessx.com', str(header_index),file_type))
        row = cursor.fetchone()
        if row is None or row[0] is None:
            raise Exception("Insert failed or no identity value returned.")
        file_id = int(row[0])
        conn.commit()
        conn.close()
        logger.info(f"Inserted file record with ID: {file_id} and header_index: {header_index}")
        return file_id
    except pyodbc.Error as e:
        logger.error(f"Database error: {e}")
        raise
    except Exception as e:
        logger.error(f"Error in insert_file_db: {e}")
        raise

# ... (Previous imports remain the same, add csv import if needed)
import csv

# Revised function to insert data into utb_nikofferloadinitial
# New function to load data into utb_nikofferloadinitial
def load_nikoffer_db(rows, file_id, headers=None, logger=None):
    logger = logger or default_logger
    try:
        file_id = int(file_id)
        with get_db_connection() as connection:
            cursor = connection.cursor()
            cursor.execute("SELECT COUNT(*) FROM utb_ImageScraperFiles WHERE ID = ?", (file_id,))
            if cursor.fetchone()[0] == 0:
                raise ValueError(f"FileID {file_id} not found in utb_ImageScraperFiles")

            # Define columns for utb_nikofferloadinitial
            nik_columns = ['FileID'] + [f'f{i}' for i in range(41)]  # FileID + f0 to f40

            df = pd.DataFrame(rows)
            logger.debug(f"Raw DataFrame for nikoffer (rows={len(df)}): {df.to_dict(orient='records')}")

            # Use provided headers or DataFrame columns
            file_columns = headers if headers else df.columns.tolist()
            logger.debug(f"File columns: {file_columns}")

            # Map file columns to f0-f40 (up to 41), pad with None if fewer, truncate if more
            mapped_columns = file_columns[:41] if len(file_columns) >= 41 else file_columns + [None] * (41 - len(file_columns))

            # Insert rows into utb_nikofferloadinitial
            rows_inserted = 0
            for idx, row in df.iterrows():
                try:
                    # Prepare values: FileID + up to 41 columns, convert to strings to keep original form
                    row_values = [file_id]
                    for col in mapped_columns:
                        if col is None:
                            row_values.append(None)
                        else:
                            value = row.get(col, None)
                            row_values.append(str(value) if value is not None else None)
                    
                    cursor.execute(
                        f"INSERT INTO utb_nikofferloadinitial ({', '.join(nik_columns)}) "
                        f"VALUES ({', '.join(['?'] * len(nik_columns))})",
                        tuple(row_values)
                    )
                    rows_inserted += 1
                except Exception as e:
                    logger.error(f"Error inserting row {idx + 1} into utb_nikofferloadinitial: {e}")

            connection.commit()
            logger.info(f"Committed {rows_inserted} rows into utb_nikofferloadinitial for FileID: {file_id}")

            # Verify insertion
            cursor.execute(
                f"SELECT FileID, {', '.join([f'f{i}' for i in range(41)])} "
                f"FROM utb_nikofferloadinitial WHERE FileID = ? ORDER BY FileID",
                (file_id,)
            )
            inserted_rows = cursor.fetchall()
            logger.debug(f"All data from utb_nikofferloadinitial for FileID {file_id}: {inserted_rows}")

        logger.info(f"Loaded {len(df)} rows into utb_nikofferloadinitial for FileID: {file_id}")
        return df
    except Exception as e:
        logger.error(f"Error loading nikoffer data: {e}")
        raise

# Updated /submitFullFile endpoint for dynamic file uploads
@app.post("/submitFullFile")
async def submit_full_file(
    fileUpload: UploadFile,
    header_index: int = Form(...),
    sendToEmail: Optional[str] = Form(None),
):
    temp_dir = None
    extracted_images_dir = None
    try:
        file_id = str(uuid.uuid4())
        default_logger.info(f"Processing full file for FileID: {file_id}")
        default_logger.info(f"Received: header_index={header_index}, sendToEmail={sendToEmail}")

        if header_index < 1:
            raise HTTPException(status_code=400, detail="header_index must be 1 or greater (1-based row number)")

        # Create temporary directory for file processing
        temp_dir = os.path.join("temp_files", "full_files", file_id)
        os.makedirs(temp_dir, exist_ok=True)
        uploaded_file_path = os.path.join(temp_dir, fileUpload.filename)
        with open(uploaded_file_path, "wb") as buffer:
            shutil.copyfileobj(fileUpload.file, buffer)

        upload_timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        # Upload file to S3 and R2
        s3_key_excel = f"luxurymarket/supplier/offer/{upload_timestamp}/{file_id}/{fileUpload.filename}"
        urls = upload_to_s3(
            uploaded_file_path,
            S3_CONFIG['bucket_name'],
            s3_key_excel,
            r2_bucket_name=S3_CONFIG['r2_bucket_name'],
            logger=default_logger,
            file_id=file_id
        )
        file_url_s3 = urls['s3']  # S3 URL for database
        file_url_r2 = urls.get('r2')  # Public R2 URL for response
        default_logger.info(f"File uploaded to S3: {file_url_s3}")
        if file_url_r2:
            default_logger.info(f"File also uploaded to R2: {file_url_r2}")

        # Determine file type and extract data
        extracted_data = []
        headers = []
        file_extension = os.path.splitext(fileUpload.filename)[1].lower()

        if file_extension in ['.xlsx', '.xls']:
            # Process Excel file
            wb = load_workbook(uploaded_file_path)
            sheet = wb.active
            image_loader = SheetImageLoader(sheet)

            default_logger.info(f"Processing Excel file with header row: {header_index}, max_row: {sheet.max_row}")

            # Get headers
            for col in sheet[header_index]:
                headers.append(str(col.value) if col.value else f"Column_{col.column}")
            default_logger.debug(f"Headers: {headers}")

            # Extract data and images
            for row_idx in range(header_index + 1, sheet.max_row + 1):
                default_logger.debug(f"Processing row {row_idx}")
                row_values = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, len(headers) + 1)]
                if all(val is None for val in row_values):
                    default_logger.info(f"Skipping empty row {row_idx}")
                    continue

                data = {header: value for header, value in zip(headers, row_values)}
                
                # Check for images in first column (Picture)
                cell_ref = f"A{row_idx}"
                extracted_images_dir = os.path.join("temp_files", "extracted_images", file_id)
                os.makedirs(extracted_images_dir, exist_ok=True)
                if image_loader.image_in(cell_ref):
                    img_path = os.path.join(extracted_images_dir, f"image_{file_id}_{cell_ref}.png")
                    image = image_loader.get(cell_ref)
                    if image:
                        image.save(img_path)
                        s3_key = f"images/{file_id}/{os.path.basename(img_path)}"
                        img_urls = upload_to_s3(
                            img_path,
                            S3_CONFIG['bucket_name'],
                            s3_key,
                            r2_bucket_name=S3_CONFIG['r2_bucket_name'],
                            logger=default_logger,
                            file_id=file_id
                        )
                        image_col = next((h for h in headers if 'image' in h.lower()), 'ExcelRowImageRef')
                        data[image_col] = img_urls['s3']
                        default_logger.info(f"Extracted and uploaded image from {cell_ref} to S3: {img_urls['s3']}")
                        if 'r2' in img_urls:
                            default_logger.info(f"Image also uploaded to R2: {img_urls['r2']}")

                extracted_data.append(data)
                default_logger.info(f"Extracted data for row {row_idx}: {data}")

        elif file_extension == '.csv':
            # Process CSV file
            df = pd.read_csv(uploaded_file_path, header=header_index - 1, encoding='utf-8', low_memory=False)
            df = df.where(pd.notnull(df), None)  # Convert NaN to None
            headers = [str(col) if col else f"Column_{i+1}" for i, col in enumerate(df.columns)]
            default_logger.debug(f"Headers: {headers}")

            for idx, row in df.iterrows():
                if all(val is None for val in row):
                    default_logger.info(f"Skipping empty row {idx + header_index}")
                    continue
                data = {header: row[col] for header, col in zip(headers, df.columns)}
                extracted_data.append(data)
                default_logger.info(f"Extracted data for row {idx + header_index}: {data}")

        else:
            raise HTTPException(status_code=400, detail=f"Unsupported file type: {file_extension}")

        default_logger.info(f"Total rows extracted (excluding header): {len(extracted_data)}")
        default_logger.info(f"Extracted for email: {sendToEmail}")

        # Insert file metadata into database
        file_id_db = insert_file_db(fileUpload.filename, file_url_s3, sendToEmail, header_index, 2, default_logger)

        # Load extracted data into utb_ImageScraperRecords
        column_map = infer_column_map(extracted_data)
        load_payload_db(extracted_data, file_id_db, column_map, default_logger)

        # Load extracted data into utb_nikofferloadinitial
        load_nikoffer_db(extracted_data, file_id_db, headers, default_logger)
        try:
            await send_file_details_email(
                to_email=sendToEmail or "nik@luxurymarket.com",
                file_id=file_id_db,
                filename=fileUpload.filename,
                s3_url=file_url_s3,
                r2_url=file_url_r2,
                record_count=len(extracted_data),
                nikoffer_count=len(extracted_data),  # Assuming same count for nikoffer
                user_email=sendToEmail or "nik@accessx.com"
            )
        except Exception as e:
            default_logger.error(f"Failed to send notification email: {e}")
            # Optionally continue despite email failure
            pass
        return {
            "success": True,
            "s3_url": file_url_s3,
            "r2_url": file_url_r2,
            "message": "File uploaded and processed successfully",
            "file_id": file_id_db
        }
    except Exception as e:
        default_logger.error(f"Error processing file: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Server error: {str(e)}")
    finally:
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)
        if extracted_images_dir and os.path.exists(extracted_images_dir):
            shutil.rmtree(extracted_images_dir, ignore_errors=True)


def infer_column_map(extracted_data: List[Dict]) -> Dict[str, str]:
    """
    Infer column mappings based on common column names in the extracted data.
    Returns a column map compatible with load_payload_db.
    """
    if not extracted_data:
        return {'brand': None, 'style': None, 'image': None, 'color': None, 'category': None}

    # Get the first row to inspect available columns
    columns = list(extracted_data[0].keys())
    default_logger.debug(f"Inferring column map from columns: {columns}")

    # Common patterns for column names
    patterns = {
        'style': [r'model', r'style', r'product_model', r'sku', r'item'],
        'brand': [r'brand', r'manufacturer', r'product_brand'],
        'color': [r'color', r'colour', r'product_color'],
        'category': [r'category', r'product_category', r'type'],
        'image': [r'image', r'photo', r'picture', r'excelrowimageref']
    }

    column_map = {'brand': None, 'style': None, 'image': None, 'color': None, 'category': None}
    
    for col in columns:
        col_lower = col.lower()
        for field, field_patterns in patterns.items():
            if any(re.search(pattern, col_lower) for pattern in field_patterns):
                if not column_map[field]:  # Only assign if not already set
                    column_map[field] = col
                    default_logger.debug(f"Mapped column '{col}' to field '{field}'")

    default_logger.info(f"Inferred column map: {column_map}")
    return column_map

def extract_full_data_and_images(
    file_path: str,
    file_id: str,
    header_row: int
) -> Tuple[List[Dict], Optional[str]]:
    """
    Extract all data and images from an Excel file without specific column targets.
    """
    wb = load_workbook(file_path)
    sheet = wb.active
    image_loader = SheetImageLoader(sheet)  # Always check for images

    extracted_images_dir = os.path.join("temp_files", "extracted_images", file_id)
    os.makedirs(extracted_images_dir, exist_ok=True)

    header_idx = header_row
    default_logger.info(f"Processing Excel file with header row: {header_idx}, max_row: {sheet.max_row}")

    # Get headers
    headers = []
    for col in sheet[header_idx]:
        if col.value:
            headers.append(str(col.value))
        else:
            headers.append(f"Column_{col.column}")  # Fallback for empty headers
    default_logger.debug(f"Headers: {headers}")

    extracted_data = []
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        default_logger.debug(f"Processing row {row_idx}")
        # Check if row is empty
        row_values = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, len(headers) + 1)]
        if all(val is None for val in row_values):
            default_logger.info(f"Skipping empty row {row_idx}")
            continue

        data = {header: None for header in headers}
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell_ref = f"{cell.column_letter}{row_idx}"
            data[header] = cell.value

            # Check for images in the cell
            if image_loader.image_in(cell_ref):
                img_path = os.path.join(extracted_images_dir, f"image_{file_id}_{cell_ref}.png")
                image = image_loader.get(cell_ref)
                if image:
                    image.save(img_path)
                    s3_key = f"images/{file_id}/{os.path.basename(img_path)}"
                    urls = upload_to_s3(
                        img_path,
                        S3_CONFIG['bucket_name'],
                        s3_key,
                        r2_bucket_name=S3_CONFIG['r2_bucket_name'],
                        logger=default_logger,
                        file_id=file_id
                    )
                    # Store image URL in a dedicated column if not already present
                    image_col = next((h for h in headers if 'image' in h.lower()), 'ExcelRowImageRef')
                    if image_col not in data:
                        data[image_col] = urls['s3']
                    else:
                        data[image_col] = urls['s3']
                    default_logger.info(f"Extracted and uploaded image from {cell_ref} to S3: {urls['s3']}")
                    if 'r2' in urls:
                        default_logger.info(f"Image also uploaded to R2: {urls['r2']}")

        extracted_data.append(data)
        default_logger.info(f"Extracted data for row {row_idx}: {data}")

    default_logger.info(f"Total rows extracted (excluding header): {len(extracted_data)}")
    return extracted_data, extracted_images_dir
# Updated /submitImage endpoint
@app.post("/submitImage")
async def submit_image(
    fileUploadImage: UploadFile,
    header_index: int = Form(...),
    imageColumnImage: Optional[str] = Form(None),
    searchColImage: str = Form(...),
    brandColImage: str = Form(...),
    ColorColImage: Optional[str] = Form(None),
    CategoryColImage: Optional[str] = Form(None),
    sendToEmail: Optional[str] = Form(None),
    manualBrand: Optional[str] = Form(None),
    isIconDistro: bool = Form(...)
):
    temp_dir = None
    extracted_images_dir = None
    try:
        file_id = str(uuid.uuid4())
        default_logger.info(f"Processing file for FileID: {file_id}")
        default_logger.info(f"Received: brandColImage={brandColImage}, manualBrand={manualBrand}, "
                           f"searchColImage={searchColImage}, imageColumnImage={imageColumnImage}, "
                           f"ColorColImage={ColorColImage}, CategoryColImage={CategoryColImage}, "
                           f"header_index={header_index}")

        if header_index < 1:
            raise HTTPException(status_code=400, detail="header_index must be 1 or greater (1-based row number)")

        temp_dir = os.path.join("temp_files", "images", file_id)
        os.makedirs(temp_dir, exist_ok=True)
        uploaded_file_path = os.path.join(temp_dir, fileUploadImage.filename)
        with open(uploaded_file_path, "wb") as buffer:
            shutil.copyfileobj(fileUploadImage.file, buffer)

        s3_key_excel = f"uploads/{file_id}/{fileUploadImage.filename}"
        urls = upload_to_s3(
            uploaded_file_path, 
            S3_CONFIG['bucket_name'], 
            s3_key_excel, 
            r2_bucket_name=S3_CONFIG['r2_bucket_name'],
            logger=default_logger,
            file_id=file_id
        )
        file_url_s3 = urls['s3']  # S3 URL for database
        file_url_r2 = urls.get('r2')  # Public R2 URL for response
        default_logger.info(f"Excel file uploaded to S3: {file_url_s3}")
        if file_url_r2:
            default_logger.info(f"Excel file also uploaded to R2: {file_url_r2}")

        extract_column_map = {
            'brand': 'MANUAL' if brandColImage == 'MANUAL' else validate_column(brandColImage),
            'style': validate_column(searchColImage),
            'image': validate_column(imageColumnImage) if imageColumnImage else None,
            'color': validate_column(ColorColImage) if ColorColImage else None,
            'category': validate_column(CategoryColImage) if CategoryColImage else None,
            'manualBrand': manualBrand,
            'isIconDistro':isIconDistro,
        }
        default_logger.info(f"Column map: {extract_column_map}")

        if extract_column_map['brand'] == 'MANUAL' and not manualBrand:
            raise HTTPException(status_code=400, detail="manualBrand is required when brandColImage is 'MANUAL'")
        data_start_row = header_index + 1
        extracted_data, extracted_images_dir = extract_data_and_images(
            uploaded_file_path, file_id, extract_column_map, data_start_row,
            manualBrand if extract_column_map['brand'] == 'MANUAL' else None
        )
        default_logger.debug(f"Extracted data: {extracted_data}")
        default_logger.info(f"Extracted for email: {sendToEmail}")
        if isIconDistro:
            file_type = 3  # Icon distribution file type
        else:
            file_type = 1

        file_id_db = insert_file_db(fileUploadImage.filename, file_url_s3, sendToEmail, header_index, file_type, default_logger)

        load_payload_db(extracted_data, file_id_db, extract_column_map, default_logger)
        try:
            await send_file_details_email(
                to_email=sendToEmail or "nik@luxurymarket.com",
                file_id=file_id_db,
                filename=fileUploadImage.filename,
                s3_url=file_url_s3,
                r2_url=file_url_r2,
                record_count=len(extracted_data),
                nikoffer_count=len(extracted_data),  # Assuming same count for nikoffer
                user_email=sendToEmail or "nik@accessx.com"
            )
        except Exception as e:
            default_logger.error(f"Failed to send notification email: {e}")
            # Optionally continue despite email failure
            pass
        return {
            "success": True,
            "s3_url": file_url_s3,
            "r2_url": file_url_r2,
            "message": "File uploaded and processed successfully",
            "file_id": file_id_db
        }
    except Exception as e:
        default_logger.error(f"Error processing file: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Server error: {str(e)}")
    finally:
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)
        if extracted_images_dir and os.path.exists(extracted_images_dir):
            shutil.rmtree(extracted_images_dir, ignore_errors=True)

def extract_data_and_images(
    file_path: str, 
    file_id: str, 
    column_map: Dict[str, str], 
    header_row: int, 
    manual_brand: Optional[str] = None
) -> Tuple[List[Dict], Optional[str]]:
    wb = load_workbook(file_path)
    sheet = wb.active
    image_loader = SheetImageLoader(sheet) if column_map.get('image') else None

    extracted_images_dir = None
    if column_map.get('image'):
        extracted_images_dir = os.path.join("temp_files", "extracted_images", file_id)
        os.makedirs(extracted_images_dir, exist_ok=True)

    header_idx = header_row 
    default_logger.info(f"Processing Excel file with header row: {header_idx}, max_row: {sheet.max_row}")

    # Define valid columns, excluding 'MANUAL' and manualBrand
    valid_columns = [col for col in column_map.values() if col and col != 'MANUAL' and col != manual_brand]
    default_logger.debug(f"Valid columns: {valid_columns}")

    header_data = {
        'search': sheet[f'{column_map["style"]}{header_idx}'].value if column_map.get('style') else None,
        'brand': manual_brand if column_map.get('brand') == 'MANUAL' else (
            sheet[f'{column_map["brand"]}{header_idx}'].value if column_map.get('brand') else None
        ),
        'color': sheet[f'{column_map["color"]}{header_idx}'].value if column_map.get('color') else None,
        'category': sheet[f'{column_map["category"]}{header_idx}'].value if column_map.get('category') else None,
    }
    default_logger.info(f"Header row {header_idx} data: {header_data}")

    extracted_data = []
    for row_idx in range(header_row + 1 , sheet.max_row + 1):
        default_logger.debug(f"Processing row {row_idx}")
        # Skip rows where all specified columns are empty
        row_is_empty = False
        if valid_columns:
            try:
                cell_values = [sheet[f'{col}{row_idx}'].value for col in valid_columns]
                row_is_empty = all(val is None for val in cell_values)
                default_logger.debug(f"Row {row_idx} cell values: {dict(zip(valid_columns, cell_values))}")
            except ValueError as e:
                default_logger.error(f"Invalid cell reference in row {row_idx}: {str(e)}")
                raise HTTPException(status_code=400, detail=f"Invalid cell reference in row {row_idx}: {str(e)}")
        
        if row_is_empty:
            default_logger.info(f"Skipping empty row {row_idx}")
            continue

        image_ref = None
        if column_map.get('image'):
            image_cell = f'{column_map["image"]}{row_idx}'
            try:
                image_ref = sheet[image_cell].value if sheet[image_cell] else None
                default_logger.debug(f"Image cell {image_cell} value: {image_ref}")
            except ValueError:
                default_logger.error(f"Invalid image cell reference: {image_cell}")
                image_ref = None

        brand = manual_brand if column_map['brand'] == 'MANUAL' else (
            sheet[f'{column_map["brand"]}{row_idx}'].value 
            if column_map.get('brand') and column_map['brand'] != 'MANUAL' else None
        )

        data = {
            'search': (
                str(sheet[f'{column_map["style"]}{row_idx}'].value) 
                if column_map.get('style') else None
            ),
            'brand': str(brand) if brand is not None else None,
            'ExcelRowImageRef': image_ref,
            'color': (
                str(sheet[f'{column_map["color"]}{row_idx}'].value) 
                if column_map.get('color') else None
            ),
            'category': (
                str(sheet[f'{column_map["category"]}{row_idx}'].value) 
                if column_map.get('category') else None
            ),
        }
        default_logger.info(f"Extracted data for row {row_idx}: {data}")

        if column_map.get('image') and not data['ExcelRowImageRef'] and image_loader and image_loader.image_in(image_cell):
            img_path = os.path.join(extracted_images_dir, f"image_{file_id}_{image_cell}.png")
            image = image_loader.get(image_cell)
            if image:
                image.save(img_path)
                s3_key = f"images/{file_id}/{os.path.basename(img_path)}"
                urls = upload_to_s3(
                    img_path, 
                    S3_CONFIG['bucket_name'], 
                    s3_key,
                    r2_bucket_name=S3_CONFIG['r2_bucket_name'],
                    logger=default_logger,
                    file_id=file_id
                )
                data['ExcelRowImageRef'] = urls['s3']  # Use S3 URL for database
                default_logger.info(f"Extracted and uploaded image from {image_cell} to S3: {urls['s3']}")
                if 'r2' in urls:
                    default_logger.info(f"Image also uploaded to R2: {urls['r2']}")
            else:
                default_logger.warning(f"No image retrieved from cell {image_cell}")

        extracted_data.append(data)

    default_logger.info(f"Total rows extracted (excluding header): {len(extracted_data)}")
    return extracted_data, extracted_images_dir

@app.api_route("/api/update-references", methods=["GET", "POST"], response_model=ReferenceData)
async def update_references(updated_data: Optional[ReferenceData] = None):
    github_url = "https://raw.githubusercontent.com/iconluxurygroup/settings-static-data/refs/heads/main/optimal-references.json"
    s3_key = "optimal-references.json"

    if updated_data is None:  # GET request
        try:
            response = requests.get(github_url)
            if response.status_code != 200:
                raise HTTPException(status_code=500, detail="Failed to fetch data from GitHub")
            data = response.json()
            return ReferenceData(data=data)
        except Exception as e:
            default_logger.error(f"Error fetching from GitHub: {e}")
            raise HTTPException(status_code=500, detail=f"Error fetching data: {str(e)}")

    else:  # POST request
        try:
            if not updated_data.data:
                raise HTTPException(status_code=400, detail="Data cannot be empty")
            if len(updated_data.data) != len(set(updated_data.data.keys())):
                raise HTTPException(status_code=400, detail="Duplicate categories are not allowed")
            if any(not key or not value for key, value in updated_data.data.items()):
                raise HTTPException(status_code=400, detail="All fields must be non-empty")

            temp_file = "temp_references.json"
            with open(temp_file, "w") as f:
                json.dump(updated_data.data, f)

            urls = upload_to_s3(
                temp_file,
                S3_CONFIG["bucket_name"],
                s3_key,
                r2_bucket_name=S3_CONFIG['r2_bucket_name'],
                logger=default_logger
            )
            s3_url = urls['s3']
            default_logger.info(f"Uploaded updated references to S3: {s3_url}")
            if 'r2' in urls:
                default_logger.info(f"References also uploaded to R2: {urls['r2']}")

            os.remove(temp_file)

            return ReferenceData(data=updated_data.data)
        except Exception as e:
            default_logger.error(f"Error uploading to S3/R2: {e}")
            raise HTTPException(status_code=500, detail=f"Error saving data: {str(e)}")

@app.get("/api/scraping-jobs", response_model=list[JobSummary])
async def get_all_jobs(page: int = 1, page_size: int = 10):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        offset = (page - 1) * page_size

        query = """
            SELECT 
                ID, 
                FileName, 
                CreateFileCompleteTime, 
                UserEmail,
                (SELECT COUNT(*) FROM utb_ImageScraperRecords WHERE FileID = utb_ImageScraperFiles.ID) as rec_count,
                (SELECT COUNT(*) FROM utb_ImageScraperResult 
                 WHERE EntryID IN (SELECT EntryID FROM utb_ImageScraperRecords WHERE FileID = utb_ImageScraperFiles.ID)) as img_count
            FROM utb_ImageScraperFiles
            WHERE FileTypeID = 1
            ORDER BY ID DESC
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """
        cursor.execute(query, (offset, page_size))
        rows = cursor.fetchall()

        jobs_data = [
            {
                "id": row.ID,
                "inputFile": row.FileName,
                "fileEnd": row.CreateFileCompleteTime.isoformat() if row.CreateFileCompleteTime else None,
                "user": row.UserEmail or "Unknown",
                "rec": row.rec_count,
                "img": row.img_count,
            }
            for row in rows
        ]

        conn.close()
        return jobs_data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/scraping-jobs/{job_id}", response_model=JobDetails)
async def get_job(job_id: int):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        query_files = """
            SELECT ID, FileName, ImageStartTime, CreateFileStartTime, CreateFileCompleteTime,
                   FileLocationURLComplete AS ResultFile, FileLocationUrl, LogFileURL, UserEmail,
                   ImageCompleteTime
            FROM utb_ImageScraperFiles
            WHERE ID = ?
        """
        cursor.execute(query_files, (job_id,))
        row = cursor.fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Job not found")

        query_results = """
            SELECT ResultID, EntryID, ImageUrl, ImageDesc, ImageSource, CreateTime, ImageUrlThumbnail,
                   SortOrder, ImageIsFashion, AiCaption, AiJson, AiLabel
            FROM utb_ImageScraperResult
            WHERE EntryID IN (SELECT EntryID FROM utb_ImageScraperRecords WHERE FileID = ?)
        """
        cursor.execute(query_results, (job_id,))
        results = cursor.fetchall()

        query_records = """
            SELECT EntryID, FileID, ExcelRowID, ProductModel, ProductBrand, CreateTime, Step1, Step2, 
                   Step3, Step4, CompleteTime, ProductColor, ProductCategory, excelRowImageRef
            FROM utb_ImageScraperRecords
            WHERE FileID = ?
        """
        cursor.execute(query_records, (job_id,))
        records = cursor.fetchall()

        job_data = {
            "id": row.ID,
            "inputFile": row.FileName,
            "imageStart": row.ImageStartTime.isoformat() if row.ImageStartTime else None,
            "fileStart": row.CreateFileStartTime.isoformat() if row.CreateFileStartTime else None,
            "fileEnd": row.CreateFileCompleteTime.isoformat() if row.CreateFileCompleteTime else None,
            "resultFile": row.ResultFile,
            "fileLocationUrl": row.FileLocationUrl,
            "logFileUrl": row.LogFileURL,
            "user": row.UserEmail or "Unknown",
            "rec": len(records),
            "img": len(results),
            "apiUsed": "google-serp",
            "imageEnd": row.ImageCompleteTime.isoformat() if row.ImageCompleteTime else None,
            "results": [
                {
                    "resultId": r.ResultID,
                    "entryId": r.EntryID,
                    "imageUrl": r.ImageUrl or "None",
                    "imageDesc": r.ImageDesc,
                    "imageSource": r.ImageSource,
                    "createTime": r.CreateTime.isoformat() if r.CreateTime else None,
                    "imageUrlThumbnail": r.ImageUrlThumbnail or "None",
                    "sortOrder": r.SortOrder or -1,
                    "imageIsFashion": r.ImageIsFashion,
                    "aiCaption": r.AiCaption,
                    "aiJson": r.AiJson,
                    "aiLabel": r.AiLabel,
                } for r in results
            ],
            "records": [
                {
                    "entryId": r.EntryID,
                    "fileId": r.FileID,
                    "excelRowId": r.ExcelRowID,
                    "productModel": r.ProductModel,
                    "productBrand": r.ProductBrand,
                    "createTime": r.CreateTime.isoformat() if r.CreateTime else None,
                    "step1": r.Step1.isoformat() if r.Step1 else None,
                    "step2": r.Step2.isoformat() if r.Step2 else None,
                    "step3": r.Step3.isoformat() if r.Step3 else None,
                    "step4": r.Step4.isoformat() if r.Step4 else None,
                    "completeTime": r.CompleteTime.isoformat() if r.CompleteTime else None,
                    "productColor": r.ProductColor if r.ProductColor else None,
                    "productCategory": r.ProductCategory if r.ProductCategory else None,
                    "excelRowImageRef": r.excelRowImageRef if r.excelRowImageRef else None,
                } for r in records
            ],
        }

        conn.close()
        return job_data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/whitelist-domains", response_model=List[DomainAggregation])
async def get_whitelist_domains():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        query = """
            SELECT ImageSource, SortOrder
            FROM utb_ImageScraperResult
        """
        cursor.execute(query)
        rows = cursor.fetchall()

        domain_data = {}
        for row in rows:
            image_source = row.ImageSource if row.ImageSource else "unknown"
            sort_order = row.SortOrder if row.SortOrder is not None else -1

            try:
                from urllib.parse import urlparse
                domain = urlparse(image_source).hostname or "unknown"
                domain = domain.replace("www.", "")
            except:
                domain = "unknown"

            if domain not in domain_data:
                domain_data[domain] = {"totalResults": 0, "positiveSortOrderCount": 0}
            
            domain_data[domain]["totalResults"] += 1
            if sort_order > 0:
                domain_data[domain]["positiveSortOrderCount"] += 1

        aggregated_domains = [
            {
                "domain": domain,
                "totalResults": data["totalResults"],
                "positiveSortOrderCount": data["positiveSortOrderCount"],
            }
            for domain, data in domain_data.items()
        ]

        conn.close()
        return aggregated_domains
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# Response model for supplier offer summary
class OfferSummary(BaseModel):
    id: int
    fileName: str
    fileLocationUrl: str
    userEmail: Optional[str]
    createTime: Optional[str]
    recordCount: int
    nikOfferCount: int

# Response model for detailed offer data
class OfferDetails(BaseModel):
    id: int
    fileName: str
    fileLocationUrl: str
    userEmail: Optional[str]
    createTime: Optional[str]
    recordCount: int
    nikOfferCount: int
    sampleRecords: List[dict]
    sampleNikOffers: List[dict]

# Endpoint to list supplier offers
@app.get("/api/luxurymarket/supplier/offers", response_model=List[OfferSummary])
async def list_supplier_offers(page: int = 1, page_size: int = 10):
    """
    List supplier offers with pagination.
    Returns a list of offers with metadata and counts of associated records.
    """
    try:
        default_logger.info(f"Fetching supplier offers: page={page}, page_size={page_size}")
        
        conn = get_db_connection()
        cursor = conn.cursor()

        offset = (page - 1) * page_size

        query = """
            SELECT 
                ID, 
                FileName, 
                FileLocationUrl, 
                UserEmail, 
                CreateFileStartTime,
                (SELECT COUNT(*) FROM utb_ImageScraperRecords WHERE FileID = utb_ImageScraperFiles.ID) as record_count,
                (SELECT COUNT(*) FROM utb_nikofferloadinitial WHERE FileID = utb_ImageScraperFiles.ID) as nikoffer_count
            FROM utb_ImageScraperFiles
            WHERE FileTypeID = 2
            ORDER BY CreateFileStartTime DESC
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """
        cursor.execute(query, (offset, page_size))
        rows = cursor.fetchall()

        offers = [
            OfferSummary(
                id=row.ID,
                fileName=row.FileName,
                fileLocationUrl=row.FileLocationUrl,
                userEmail=row.UserEmail,
                createTime=row.CreateFileStartTime.isoformat() if row.CreateFileStartTime else None,
                recordCount=row.record_count,
                nikOfferCount=row.nikoffer_count
            )
            for row in rows
        ]

        default_logger.info(f"Retrieved {len(offers)} supplier offers")
        conn.close()
        return offers
    except Exception as e:
        default_logger.error(f"Error listing supplier offers: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Server error: {str(e)}")
# Add to imports if not already present
from pathlib import Path
import uuid
import urllib.parse

# New function to load data into utb_OfferImport
def load_offer_import_db(rows, file_id, logger=None):
    logger = logger or default_logger
    try:
        file_id = int(file_id)
        with get_db_connection() as connection:
            cursor = connection.cursor()
            cursor.execute("SELECT COUNT(*) FROM utb_ImageScraperFiles WHERE ID = ?", (file_id,))
            if cursor.fetchone()[0] == 0:
                raise ValueError(f"FileID {file_id} not found in utb_ImageScraperFiles")

            # Define columns for utb_OfferImport (assuming up to 41 columns like utb_nikofferloadinitial)
            offer_columns = ['FileID'] + [f'f{i}' for i in range(41)]  # FileID + f0 to f40

            # Convert rows to DataFrame for processing
            df = pd.DataFrame(rows)
            logger.debug(f"Raw DataFrame for utb_OfferImport (rows={len(df)}): {df.to_dict(orient='records')}")

            # Ensure all values are strings or None to preserve original format
            df = df.astype(str).where(df.notna(), None)

            # Insert rows into utb_OfferImport
            rows_inserted = 0
            for idx, row in df.iterrows():
                try:
                    # Prepare values: FileID + up to 41 columns
                    row_values = [file_id] + [row.get(i, None) for i in range(min(len(row), 41))]
                    # Pad with None if fewer than 41 columns
                    row_values += [None] * (42 - len(row_values))
                    
                    cursor.execute(
                        f"INSERT INTO utb_OfferImport ({', '.join(offer_columns)}) "
                        f"VALUES ({', '.join(['?'] * len(offer_columns))})",
                        tuple(row_values)
                    )
                    rows_inserted += 1
                except Exception as e:
                    logger.error(f"Error inserting row {idx + 1} into utb_OfferImport: {e}")

            connection.commit()
            logger.info(f"Committed {rows_inserted} rows into utb_OfferImport for FileID: {file_id}")

            # Verify insertion
            cursor.execute(
                f"SELECT FileID, {', '.join([f'f{i}' for i in range(41)])} "
                f"FROM utb_OfferImport WHERE FileID = ? ORDER BY FileID",
                (file_id,)
            )
            inserted_rows = cursor.fetchall()
            logger.debug(f"All data from utb_OfferImport for FileID {file_id}: {inserted_rows}")

        logger.info(f"Loaded {len(df)} rows into utb_OfferImport for FileID: {file_id}")
        return df
    except Exception as e:
        logger.error(f"Error loading offer import data: {e}")
        raise

@app.post("/submitOffer")
async def submit_offer(
    fileUrl: str = Form(...),
    header_index: int = Form(...),
    sendToEmail: Optional[str] = Form(None),
):
    temp_dir = None
    extracted_images_dir = None
    try:
        file_id = str(uuid.uuid4())
        default_logger.info(f"Processing offer file for FileID: {file_id}")
        default_logger.info(f"Received: fileUrl={fileUrl}, header_index={header_index}, sendToEmail={sendToEmail}")

        if header_index < 1:
            raise HTTPException(status_code=400, detail="header_index must be 1 or greater (1-based row number)")

        # Create temporary directory for file processing
        temp_dir = os.path.join("temp_files", "offers", file_id)
        os.makedirs(temp_dir, exist_ok=True)

        # Download the file
        response = requests.get(fileUrl)
        if response.status_code != 200:
            raise HTTPException(status_code=400, detail=f"Failed to download file from {fileUrl}")

        # Extract filename from URL
        parsed_url = urllib.parse.urlparse(fileUrl)
        filename = os.path.basename(parsed_url.path)
        if not filename:
            raise HTTPException(status_code=400, detail="Could not determine filename from URL")

        uploaded_file_path = os.path.join(temp_dir, filename)
        with open(uploaded_file_path, "wb") as f:
            f.write(response.content)
        default_logger.info(f"Downloaded file to {uploaded_file_path}")

        upload_timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        # Upload file to S3 and R2
        s3_key_excel = f"luxurymarket/supplier/offer/{upload_timestamp}/{file_id}/{filename}"
        urls = upload_to_s3(
            uploaded_file_path,
            S3_CONFIG['bucket_name'],
            s3_key_excel,
            r2_bucket_name=S3_CONFIG['r2_bucket_name'],
            logger=default_logger,
            file_id=file_id
        )
        file_url_s3 = urls['s3']  # S3 URL for database
        file_url_r2 = urls.get('r2')  # Public R2 URL for response
        default_logger.info(f"File uploaded to S3: {file_url_s3}")
        if file_url_r2:
            default_logger.info(f"File also uploaded to R2: {file_url_r2}")

        # Determine file type and process data
        extracted_data = []
        file_extension = os.path.splitext(filename)[1].lower()

        if file_extension in ['.xlsx', '.xls']:
            # Process Excel file
            wb = load_workbook(uploaded_file_path)
            sheet = wb.active

            default_logger.info(f"Processing Excel file, max_row: {sheet.max_row}")

            # Create directory for extracted images
            extracted_images_dir = os.path.join("temp_files", "extracted_images", file_id)
            os.makedirs(extracted_images_dir, exist_ok=True)

            # Read all rows starting from row 1 (no header processing)
            for row_idx in range(1, sheet.max_row + 1):
                row_values = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, sheet.max_column + 1)]
                # Skip empty rows
                if all(val is None or str(val).strip() == '' for val in row_values):
                    default_logger.info(f"Skipping empty row {row_idx}")
                    continue

                # Check for image filename in first column
                if row_values[0] and isinstance(row_values[0], str) and re.match(r'.+\.(png|jpg|jpeg|gif)$', row_values[0], re.IGNORECASE):
                    image_filename = row_values[0]
                    image_path = os.path.join(extracted_images_dir, image_filename)
                    # Assume image file exists in the same directory as the uploaded file or needs to be provided
                    # For this implementation, we'll assume the image is accessible locally or needs to be downloaded
                    # If images are in the same directory as the file, adjust the path accordingly
                    if os.path.exists(image_path):
                        s3_key = f"images/{file_id}/{image_filename}"
                        img_urls = upload_to_s3(
                            image_path,
                            S3_CONFIG['bucket_name'],
                            s3_key,
                            r2_bucket_name=S3_CONFIG['r2_bucket_name'],
                            logger=default_logger,
                            file_id=file_id
                        )
                        # Replace the image filename with the R2 URL
                        row_values[0] = img_urls.get('r2', img_urls['s3'])
                        default_logger.info(f"Uploaded image {image_filename} to R2: {row_values[0]}")
                    else:
                        default_logger.warning(f"Image file {image_filename} not found at {image_path}")

                extracted_data.append(row_values)
                default_logger.info(f"Extracted data for row {row_idx}: {row_values}")

        elif file_extension == '.csv':
            # Process CSV file
            with open(uploaded_file_path, 'r', encoding='utf-8') as csv_file:
                csv_reader = csv.reader(csv_file)
                for row_idx, row in enumerate(csv_reader, start=1):
                    # Skip empty rows
                    row_values = [val if val.strip() != '' else None for val in row]
                    if all(val is None for val in row_values):
                        default_logger.info(f"Skipping empty row {row_idx}")
                        continue

                    # Check for image filename in first column
                    if row_values[0] and re.match(r'.+\.(png|jpg|jpeg|gif)$', row_values[0], re.IGNORECASE):
                        image_filename = row_values[0]
                        image_path = os.path.join(extracted_images_dir or temp_dir, image_filename)
                        if os.path.exists(image_path):
                            s3_key = f"images/{file_id}/{image_filename}"
                            img_urls = upload_to_s3(
                                image_path,
                                S3_CONFIG['bucket_name'],
                                s3_key,
                                r2_bucket_name=S3_CONFIG['r2_bucket_name'],
                                logger=default_logger,
                                file_id=file_id
                            )
                            # Replace the image filename with the R2 URL
                            row_values[0] = img_urls.get('r2', img_urls['s3'])
                            default_logger.info(f"Uploaded image {image_filename} to R2: {row_values[0]}")
                        else:
                            default_logger.warning(f"Image file {image_filename} not found at {image_path}")

                    extracted_data.append(row_values)
                    default_logger.info(f"Extracted data for row {row_idx}: {row_values}")

        else:
            raise HTTPException(status_code=400, detail=f"Unsupported file type: {file_extension}")

        default_logger.info(f"Total rows extracted: {len(extracted_data)}")
        default_logger.info(f"Extracted for email: {sendToEmail}")

        # Insert file metadata into database
        file_id_db = insert_file_db(filename, file_url_s3, sendToEmail, header_index, 2, default_logger)

        # Load extracted data into utb_OfferImport
        load_offer_import_db(extracted_data, file_id_db, default_logger)

        # Send notification email
        try:
            await send_file_details_email(
                to_email=sendToEmail or "nik@luxurymarket.com",
                file_id=file_id_db,
                filename=filename,
                s3_url=file_url_s3,
                r2_url=file_url_r2,
                record_count=len(extracted_data),
                nikoffer_count=0,  # No nikoffer data in this case
                user_email=sendToEmail or "nik@accessx.com"
            )
        except Exception as e:
            default_logger.error(f"Failed to send notification email: {e}")
            # Continue despite email failure
            pass

        return {
            "success": True,
            "s3_url": file_url_s3,
            "r2_url": file_url_r2,
            "message": "Offer file downloaded and processed successfully",
            "file_id": file_id_db
        }

    except Exception as e:
        default_logger.error(f"Error processing offer file: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Server error: {str(e)}")

    finally:
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)
        if extracted_images_dir and os.path.exists(extracted_images_dir):
            shutil.rmtree(extracted_images_dir, ignore_errors=True)
# Endpoint to get details of a specific supplier offer
@app.get("/api/luxurymarket/supplier/offers/{offer_id}", response_model=OfferDetails)
async def get_supplier_offer(offer_id: int):
    """
    Get detailed information about a specific supplier offer, including sample records.
    """
    try:
        default_logger.info(f"Fetching details for supplier offer ID: {offer_id}")
        
        conn = get_db_connection()
        cursor = conn.cursor()

        # Fetch offer metadata
        query_offer = """
            SELECT 
                ID, 
                FileName, 
                FileLocationUrl, 
                UserEmail, 
                CreateFileStartTime,
                (SELECT COUNT(*) FROM utb_ImageScraperRecords WHERE FileID = utb_ImageScraperFiles.ID) as record_count,
                (SELECT COUNT(*) FROM utb_nikofferloadinitial WHERE FileID = utb_ImageScraperFiles.ID) as nikoffer_count
            FROM utb_ImageScraperFiles
            WHERE ID = ? AND FileTypeID = 2
        """
        cursor.execute(query_offer, (offer_id,))
        row = cursor.fetchone()

        if not row:
            default_logger.warning(f"Offer ID {offer_id} not found or not a supplier offer")
            raise HTTPException(status_code=404, detail="Supplier offer not found")

        # Fetch sample records from utb_ImageScraperRecords (limit to 5 for brevity)
        query_records = """
            SELECT TOP 5 
                EntryID, 
                ExcelRowID, 
                ProductModel, 
                ProductBrand, 
                ProductColor, 
                ProductCategory, 
                ExcelRowImageRef
            FROM utb_ImageScraperRecords
            WHERE FileID = ?
            ORDER BY ExcelRowID
        """
        cursor.execute(query_records, (offer_id,))
        records = cursor.fetchall()
        sample_records = [
            {
                "entryId": r.EntryID,
                "excelRowId": r.ExcelRowID,
                "productModel": r.ProductModel,
                "productBrand": r.ProductBrand,
                "productColor": r.ProductColor,
                "productCategory": r.ProductCategory,
                "excelRowImageRef": r.ExcelRowImageRef
            }
            for r in records
        ]

        # Fetch sample records from utb_nikofferloadinitial (limit to 5 for brevity)
        query_nikoffers = """
            SELECT TOP 5 
                FileID, 
                f0, f1, f2, f3, f4, 
                f5, f6, f7, f8, f9
            FROM utb_nikofferloadinitial
            WHERE FileID = ?
            ORDER BY FileID
        """
        cursor.execute(query_nikoffers, (offer_id,))
        nikoffers = cursor.fetchall()
        sample_nikoffers = [
            {
                "fileId": n.FileID,
                "f0": n.f0,
                "f1": n.f1,
                "f2": n.f2,
                "f3": n.f3,
                "f4": n.f4,
                "f5": n.f5,
                "f6": n.f6,
                "f7": n.f7,
                "f8": n.f8,
                "f9": n.f9
            }
            for n in nikoffers
        ]

        offer_details = OfferDetails(
            id=row.ID,
            fileName=row.FileName,
            fileLocationUrl=row.FileLocationUrl,
            userEmail=row.UserEmail,
            createTime=row.CreateFileStartTime.isoformat() if row.CreateFileStartTime else None,
            recordCount=row.record_count,
            nikOfferCount=row.nikoffer_count,
            sampleRecords=sample_records,
            sampleNikOffers=sample_nikoffers
        )

        default_logger.info(f"Retrieved details for supplier offer ID: {offer_id}")
        conn.close()
        return offer_details
    except Exception as e:
        default_logger.error(f"Error fetching supplier offer {offer_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Server error: {str(e)}")


# New endpoint to get job progress
@app.get("/api/scraping-jobs/{file_id}/progress", response_model=JobProgress)
async def get_job_progress(file_id: int):
    """
    Get the processing progress for a specific job (file).
    Calculates progress for Step1, Step2, Step3, and Step4.
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # First, check if the FileID is valid to provide a clear 404 error
        cursor.execute("SELECT COUNT(*) FROM utb_ImageScraperFiles WHERE ID = ?", (file_id,))
        if cursor.fetchone()[0] == 0:
            raise HTTPException(status_code=404, detail=f"Job with FileID {file_id} not found.")

        # Use COUNT(column) to count non-null values, which is faster than CASE statements
        query = """
            SELECT
                COUNT(*) AS TotalRecords,
                COUNT(Step1) AS Step1Completed,
                COUNT(Step2) AS Step2Completed,
                COUNT(Step3) AS Step3Completed,
                COUNT(Step4) AS Step4Completed
            FROM utb_ImageScraperRecords
            WHERE FileID = ?
        """
        cursor.execute(query, (file_id,))
        progress_data = cursor.fetchone()
        conn.close()

        if not progress_data:
            # This case is unlikely if the FileID exists, but good practice to handle
            raise HTTPException(status_code=404, detail=f"No records found for FileID {file_id}.")

        total_records = progress_data.TotalRecords
        
        # Handle the case where a file has been created but records are not yet loaded
        if total_records == 0:
            return JobProgress(
                fileId=file_id, totalRecords=0,
                step1Completed=0, step1Progress=0.0,
                step2Completed=0, step2Progress=0.0,
                step3Completed=0, step3Progress=0.0,
                step4Completed=0, step4Progress=0.0
            )

        # Calculate progress for each step
        step1_progress = (progress_data.Step1Completed / total_records) * 100
        step2_progress = (progress_data.Step2Completed / total_records) * 100
        step3_progress = (progress_data.Step3Completed / total_records) * 100
        step4_progress = (progress_data.Step4Completed / total_records) * 100

        return JobProgress(
            fileId=file_id,
            totalRecords=total_records,
            step1Completed=progress_data.Step1Completed,
            step1Progress=round(step1_progress, 2),
            step2Completed=progress_data.Step2Completed,
            step2Progress=round(step2_progress, 2),
            step3Completed=progress_data.Step3Completed,
            step3Progress=round(step3_progress, 2),
            step4Completed=progress_data.Step4Completed,
            step4Progress=round(step4_progress, 2),
        )
    except HTTPException as http_exc:
        # Re-raise known HTTP exceptions to avoid masking them as 500 errors
        raise http_exc
    except Exception as e:
        default_logger.error(f"Error getting job progress for FileID {file_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)